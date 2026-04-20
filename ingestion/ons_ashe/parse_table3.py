"""Parse ONS ASHE Table 3 Excel workbooks — Region × 2-digit SOC.

ASHE Table 3 provides the CRITICAL missing piece: salary percentiles broken
down by BOTH occupation code AND UK region.

Structure of each workbook:
    Sheets: "All", "Male", "Female", "Full-Time", "Part-Time", etc.
    We parse the "Full-Time" sheet (consistent with our Table 2 ingestion).

    Within each sheet, the data is flat/hierarchical:
        Row 6:  United Kingdom (K02000001)  ← national aggregate
        Row 7:  SOC 1 (national)
        Row 8:  SOC 11 (national)
        ...
        Row 42: North East (E12000001)      ← region header
        Row 43: SOC 1 (North East)
        Row 44: SOC 11 (North East)
        ...
        Row N:  North West (E12000002)      ← next region header
        ...

    Region headers are identified by GSS codes in column B (e.g. E12000001).
    SOC rows under each region have 1- or 2-digit numeric codes in column B.
    The description column often includes "Region, SOC description" format.

Download from:
    https://www.ons.gov.uk/employmentandlabourmarket/peopleinwork/earningsandworkinghours/datasets/regionbyoccupation2digitsocashetable3
"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

from shared.models import (
    CompensationObservation,
    ContractType,
    ExperienceBand,
    ObservationType,
    Period,
)
from shared.normalization import NORMALIZATION_VERSION, normalize_to_annual

SOURCE_ID = "ons_ashe_table3"

# Column index (1-based) → (label, observation_type, percentile)
COLUMN_MAP: dict[int, tuple[str, ObservationType, Optional[int]]] = {
    4:  ("median", ObservationType.PERCENTILE, 50),
    6:  ("mean",   ObservationType.POINT,      None),
    8:  ("p10",    ObservationType.PERCENTILE, 10),
    9:  ("p20",    ObservationType.PERCENTILE, 20),
    10: ("p25",    ObservationType.PERCENTILE, 25),
    11: ("p30",    ObservationType.PERCENTILE, 30),
    12: ("p40",    ObservationType.PERCENTILE, 40),
    13: ("p60",    ObservationType.PERCENTILE, 60),
    14: ("p70",    ObservationType.PERCENTILE, 70),
    15: ("p75",    ObservationType.PERCENTILE, 75),
    16: ("p80",    ObservationType.PERCENTILE, 80),
    17: ("p90",    ObservationType.PERCENTILE, 90),
}

# GSS codes that identify region header rows
KNOWN_GSS_CODES = {
    "K02000001",   # United Kingdom
    "E92000001",   # England
    "E12000001",   # North East
    "E12000002",   # North West
    "E12000003",   # Yorkshire and The Humber
    "E12000004",   # East Midlands
    "E12000005",   # West Midlands
    "E12000006",   # East of England
    "E12000007",   # London
    "E12000008",   # South East
    "E12000009",   # South West
    "W92000004",   # Wales
    "S92000003",   # Scotland
    "N92000002",   # Northern Ireland
}

# Map filename patterns to Period
FILE_PERIOD_MAP: list[tuple[str, Period]] = [
    ("annual pay",  Period.ANNUAL),
    ("weekly pay",  Period.WEEKLY),
    ("hourly pay",  Period.HOURLY),
]


def _detect_period(filename: str) -> Optional[Period]:
    low = filename.lower()
    for pattern, period in FILE_PERIOD_MAP:
        if pattern in low:
            return period
    return None


def _detect_year(filename: str) -> Optional[int]:
    m = re.search(r"(\d{4})\.xlsx", filename, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if s in ("", "x", "-", "..", ":", "*", "n/a", "N/A"):
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _is_gss_code(code: str) -> bool:
    """Check if a code looks like a GSS location code (e.g. E12000001, K02000001)."""
    return code in KNOWN_GSS_CODES


def _is_soc_code(code: str) -> bool:
    """Check if a code looks like a SOC occupation code (1-2 digit number)."""
    return code.isdigit() and len(code) <= 2


def parse_workbook(
    path: Path,
    sheet_name: str = "Full-Time",
) -> list[CompensationObservation]:
    """Parse a single ASHE Table 3 workbook.

    The data is flat within the sheet: region header rows (identified by GSS
    codes) are followed by SOC rows that belong to that region. We track the
    "current region" as we scan down the rows.
    """
    period = _detect_period(path.name)
    if period is None:
        raise ValueError(f"Cannot detect pay period from filename: {path.name}")

    data_year = _detect_year(path.name)
    if data_year is None:
        raise ValueError(f"Cannot detect year from filename: {path.name}")

    observed_at = date(data_year, 4, 1)

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path.name}. "
                         f"Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    observations: list[CompensationObservation] = []
    current_region: Optional[str] = None  # GSS code of current region context

    for row in ws.iter_rows(min_row=6, max_col=17, values_only=False):
        desc_cell = row[0].value
        code_cell = row[1].value

        if desc_cell is None:
            continue

        desc = str(desc_cell).strip()
        if not desc or desc.lower().startswith("note"):
            continue

        code_str = str(code_cell).strip() if code_cell is not None else ""
        # Handle float codes (e.g. 1.0 → "1")
        if code_cell is not None and isinstance(code_cell, float):
            code_str = str(int(code_cell))

        if desc.lower() == "not classified":
            continue

        # Check if this row is a region header
        if _is_gss_code(code_str):
            current_region = code_str
            # Region header rows also contain aggregate data for that region
            # (all occupations combined). We store these with occupation_code=None.
            soc_code = None
        elif _is_soc_code(code_str):
            soc_code = code_str
        else:
            # Skip rows we can't classify (e.g. blank codes, text codes)
            continue

        # Sample size from column C (thousands)
        sample_thousands = _safe_float(row[2].value)
        sample_size = int(sample_thousands * 1000) if sample_thousands else None

        for col_idx, (label, obs_type, percentile) in COLUMN_MAP.items():
            val = _safe_float(row[col_idx - 1].value)
            if val is None:
                continue

            # Sanity checks
            if period == Period.ANNUAL and val < 1000:
                continue
            if period == Period.WEEKLY and val < 20:
                continue
            if period == Period.HOURLY and val < 2:
                continue

            normalized = normalize_to_annual(val, period.value)

            # Deterministic reference for upsert idempotency
            ref = (f"ashe_t3:{data_year}:{period.value}"
                   f":{current_region or 'UK'}"
                   f":{soc_code or 'all'}:{label}")

            observations.append(CompensationObservation(
                source_id=SOURCE_ID,
                source_reference=ref,
                occupation_code=soc_code,
                location_code=current_region,
                company_ref=None,
                observation_type=obs_type,
                value_amount=val,
                value_min=None,
                value_max=None,
                percentile=percentile,
                period=period,
                normalized_annual_amount=normalized,
                normalization_method_version=NORMALIZATION_VERSION,
                currency="GBP",
                experience_band=ExperienceBand.UNKNOWN,
                contract_type=ContractType.PERMANENT,
                sample_size=sample_size,
                total_comp_annual=None,
                observed_at=observed_at,
                source_payload={
                    "table": "ASHE Table 3",
                    "sheet": sheet_name,
                    "region": current_region,
                    "year": data_year,
                    "description": desc,
                    "soc_code": soc_code,
                    "period": period.value,
                    "label": label,
                    "raw_value": val,
                    "sample_thousands": sample_thousands,
                },
            ))

    wb.close()
    return observations


def parse_directory(
    directory: Path,
    sheet_name: str = "Full-Time",
) -> list[CompensationObservation]:
    """Parse all ASHE Table 3 'a' workbooks in a directory.

    We parse the primary data files and skip CV (coefficient of variation) files.
    Target tables: 3.7a (annual), 3.1a (weekly), 3.5a (hourly) — gross pay.
    """
    all_obs: list[CompensationObservation] = []
    target_tables = ["3.7a", "3.1a", "3.5a"]

    for xlsx in sorted(directory.glob("*.xlsx")):
        name = xlsx.name
        # Skip CV files
        if " CV " in name or name.endswith("CV.xlsx"):
            continue
        # Only parse gross pay tables we want
        matched = False
        for table_id in target_tables:
            if f"Table {table_id}" in name:
                matched = True
                break
        if not matched:
            continue

        period = _detect_period(name)
        if period is None:
            continue

        print(f"[ons_ashe_t3] Parsing {name} ({period.value})...")
        try:
            obs = parse_workbook(xlsx, sheet_name)
            print(f"[ons_ashe_t3]   -> {len(obs)} observations")
            all_obs.extend(obs)
        except Exception as e:
            print(f"[ons_ashe_t3]   -> ERROR: {e}")

    return all_obs


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m ingestion.ons_ashe.parse_table3 <directory>")
        sys.exit(1)
    obs = parse_directory(Path(sys.argv[1]))
    print(f"\nTotal: {len(obs)} observations")
    if obs:
        regions = set(o.location_code for o in obs if o.location_code)
        socs = set(o.occupation_code for o in obs if o.occupation_code)
        print(f"Regions found: {len(regions)}")
        print(f"SOC codes found: {len(socs)}")
        # Count observations with BOTH occupation AND location
        both = sum(1 for o in obs if o.occupation_code and o.location_code)
        print(f"Observations with BOTH occupation + region: {both}")
        for r in sorted(regions):
            count = sum(1 for o in obs if o.location_code == r)
            print(f"  {r}: {count} observations")
