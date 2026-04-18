"""Parse ONS ASHE Table 2 Excel workbooks into CompensationObservation records.

ASHE Table 2 provides occupation-level pay data by SOC 2020 code.  The ONS
publishes these as downloadable Excel files (not via Nomis, which lacks the
occupation dimension).

We parse two workbooks:
    Table 2.7a  — Annual pay (Gross)   → period=ANNUAL, values in £/year
    Table 2.1a  — Weekly pay (Gross)   → period=WEEKLY, values in £/week

From each workbook we read the "Full-Time" sheet:
    Row 5 = headers:
        A=Description, B=Code, C=(thousand), D=Median, E=change,
        F=Mean, G=change, H=P10, I=P20, J=P25, K=P30, L=P40,
        M=P60, N=P70, O=P75, P=P80, Q=P90
    Row 6 = "All employees" (code='')  → stored with occupation_code=None
    Rows 7+ = SOC groups (code=1..92)  → stored with occupation_code=str(code)

Each cell yields one CompensationObservation of type PERCENTILE (or POINT
for the mean).
"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

from shared.models import (
    CompensationObservation,
    ContractType,
    ExperienceBand,
    ObservationType,
    Period,
)
from shared.normalization import NORMALIZATION_VERSION, normalize_to_annual

SOURCE_ID = "ons_ashe_table2"

# Column index (1-based) → (label, observation_type, percentile)
# D=4 Median, F=6 Mean, H=8 P10, I=9 P20, J=10 P25, K=11 P30,
# L=12 P40, M=13 P60, N=14 P70, O=15 P75, P=16 P80, Q=17 P90
COLUMN_MAP: dict[int, tuple[str, ObservationType, Optional[int]]] = {
    4:  ("median", ObservationType.PERCENTILE, 50),
    6:  ("mean",   ObservationType.POINT,      None),
    8:  ("p10",    ObservationType.PERCENTILE, 10),
    9:  ("p20",    ObservationType.PERCENTILE, 20),
    10: ("p25",    ObservationType.PERCENTILE, 25),
    11: ("p30",    ObservationType.PERCENTILE, 30),
    12: ("p40",    ObservationType.PERCENTILE, 40),
    13: ("p60",    ObservationType.PERCENTILE, 60),
    14: ("p70",    ObservationType.PERCENTILE, 70),
    15: ("p75",    ObservationType.PERCENTILE, 75),
    16: ("p80",    ObservationType.PERCENTILE, 80),
    17: ("p90",    ObservationType.PERCENTILE, 90),
}

# Map filename patterns to Period
FILE_PERIOD_MAP: list[tuple[str, Period]] = [
    ("annual pay",  Period.ANNUAL),
    ("weekly pay",  Period.WEEKLY),
    ("hourly pay",  Period.HOURLY),
]


def _detect_period(filename: str) -> Optional[Period]:
    low = filename.lower()
    for pattern, period in FILE_PERIOD_MAP:
        if pattern in low:
            return period
    return None


def _detect_year(filename: str) -> Optional[int]:
    """Extract the data year from the filename, e.g. '...2025.xlsx' → 2025."""
    m = re.search(r"(\d{4})\.xlsx", filename, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def _safe_float(val) -> Optional[float]:
    """Convert a cell value to float, returning None for suppressed/missing."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if s in ("", "x", "-", "..", ":", "*", "n/a", "N/A"):
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def parse_workbook(
    path: Path,
    sheet_name: str = "Full-Time",
) -> list[CompensationObservation]:
    """Parse a single ASHE Table 2 workbook into observations.

    Returns a list of CompensationObservation objects, one per
    (SOC code × percentile/mean) cell.
    """
    period = _detect_period(path.name)
    if period is None:
        raise ValueError(f"Cannot detect pay period from filename: {path.name}")

    data_year = _detect_year(path.name)
    if data_year is None:
        raise ValueError(f"Cannot detect year from filename: {path.name}")

    # ASHE reference date is April of the survey year
    observed_at = date(data_year, 4, 1)

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path.name}. "
                         f"Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    observations: list[CompensationObservation] = []

    for row in ws.iter_rows(min_row=6, max_col=17, values_only=False):
        # Column A = description, Column B = SOC code
        desc_cell = row[0].value
        code_cell = row[1].value

        if desc_cell is None:
            continue  # skip blank rows

        desc = str(desc_cell).strip()
        if not desc or desc.lower().startswith("note"):
            continue

        # SOC code: can be int (1, 11, 21...) or empty string for "All"
        soc_code: Optional[str] = None
        if code_cell is not None and str(code_cell).strip():
            soc_code = str(int(code_cell)) if isinstance(code_cell, float) else str(code_cell).strip()

        # "Not Classified" row — skip
        if desc.lower() == "not classified":
            continue

        # Sample size from column C (thousands)
        sample_thousands = _safe_float(row[2].value)
        sample_size = int(sample_thousands * 1000) if sample_thousands else None

        for col_idx, (label, obs_type, percentile) in COLUMN_MAP.items():
            val = _safe_float(row[col_idx - 1].value)  # row is 0-indexed
            if val is None:
                continue

            # Sanity: annual pay should be > £1000, weekly > £20
            if period == Period.ANNUAL and val < 1000:
                continue
            if period == Period.WEEKLY and val < 20:
                continue
            if period == Period.HOURLY and val < 2:
                continue

            normalized = normalize_to_annual(val, period.value)

            # Deterministic reference for upsert idempotency
            ref = f"ashe_t2:{data_year}:{period.value}:{soc_code or 'all'}:{label}"

            observations.append(CompensationObservation(
                source_id=SOURCE_ID,
                source_reference=ref,
                occupation_code=soc_code,
                location_code=None,  # ASHE Table 2 is national (UK-wide)
                company_ref=None,
                observation_type=obs_type,
                value_amount=val,
                value_min=None,
                value_max=None,
                percentile=percentile,
                period=period,
                normalized_annual_amount=normalized,
                normalization_method_version=NORMALIZATION_VERSION,
                currency="GBP",
                experience_band=ExperienceBand.UNKNOWN,
                contract_type=ContractType.PERMANENT,
                sample_size=sample_size,
                total_comp_annual=None,
                observed_at=observed_at,
                source_payload={
                    "table": "ASHE Table 2",
                    "sheet": sheet_name,
                    "year": data_year,
                    "description": desc,
                    "soc_code": soc_code,
                    "period": period.value,
                    "label": label,
                    "raw_value": val,
                    "sample_thousands": sample_thousands,
                },
            ))

    wb.close()
    return observations


def parse_directory(
    directory: Path,
    sheet_name: str = "Full-Time",
) -> list[CompensationObservation]:
    """Parse all ASHE Table 2 'a' workbooks (not CV files) in a directory.

    We only parse the primary data files (Table 2.Xa) and skip:
      - CV files (Table 2.Xb) — coefficient of variation, not pay data
      - Non-pay files (hours worked, overtime pay, incentive pay)
    """
    all_obs: list[CompensationObservation] = []
    target_tables = ["2.7a", "2.1a", "2.5a"]  # annual, weekly, hourly gross

    for xlsx in sorted(directory.glob("*.xlsx")):
        name = xlsx.name
        # Skip CV files
        if " CV " in name or name.endswith("CV.xlsx"):
            continue
        # Only parse gross pay tables we want
        matched = False
        for table_id in target_tables:
            if f"Table {table_id}" in name:
                matched = True
                break
        if not matched:
            continue

        period = _detect_period(name)
        if period is None:
            continue

        print(f"[ons_ashe] Parsing {name} ({period.value})...")
        try:
            obs = parse_workbook(xlsx, sheet_name)
            print(f"[ons_ashe]   → {len(obs)} observations")
            all_obs.extend(obs)
        except Exception as e:
            print(f"[ons_ashe]   → ERROR: {e}")

    return all_obs


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python -m ingestion.ons_ashe.parse <directory>")
        sys.exit(1)
    obs = parse_directory(Path(sys.argv[1]))
    print(f"\nTotal: {len(obs)} observations")
    if obs:
        print("Sample:", obs[0].to_dict())
