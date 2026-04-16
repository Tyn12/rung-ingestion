"""Parse ONS / HMRC PAYE RTI XLSX workbooks into CompensationObservation rows.

Each workbook has ~39 sheets. The structure per data sheet is:
    Row 0-4: Title, subtitle, blanks, units
    Row 5:   (optional) Code row — NUTS codes (UKC, UKD…) or SIC letters (A, B…)
    Row 6:   Label row — region names, industry names, age bands
    Row 7+:  Data rows — Column A is "July 2014", "August 2014", etc.
                         Remaining columns are numeric values (£ per month)

Sheets named "Median pay …" and "Mean pay …" are the ones we ingest.
We skip employee counts, aggregate pay, and flow sheets.

Values are monthly median/mean pay in £. We store monthly and annualise × 12.
"""
from __future__ import annotations
import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

from shared.models import (
    CompensationObservation,
    ContractType,
    ExperienceBand,
    ObservationType,
    Period,
)
from shared.normalization import NORMALIZATION_VERSION

_MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_date_cell(cell_value) -> Optional[date]:
    """Parse 'July 2014' or 'March 2026' etc. into a date."""
    if isinstance(cell_value, date):
        return cell_value
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    m = re.match(r"([A-Za-z]+)\s+(\d{4})$", s)
    if m:
        month_name, year = m.groups()
        month_num = _MONTH_NAMES.get(month_name.lower())
        if month_num:
            try:
                return date(int(year), month_num, 1)
            except ValueError:
                pass
    return None


def _classify_sheet(name: str) -> Optional[dict]:
    """Determine what kind of data a sheet contains.

    Returns None for sheets we don't ingest (employees, aggregate, flows).
    """
    low = name.lower()

    # Only ingest median and mean pay sheets
    if "median pay" not in low and "mean pay" not in low:
        return None

    stat = "median" if "median" in low else "mean"

    axis = "unknown"
    if "nuts1" in low and "age" in low:
        axis = "nuts1_age"
    elif "nuts1" in low and "sector" in low:
        axis = "nuts1_sector"
    elif "nuts1" in low:
        axis = "nuts1"
    elif "nuts2" in low:
        axis = "nuts2"
    elif "nuts3" in low:
        axis = "nuts3"
    elif "la" in low:
        axis = "local_authority"
    elif "industry" in low:
        axis = "industry"
    elif "age" in low:
        axis = "age"
    elif "uk" in low:
        axis = "uk"

    return {"stat": stat, "axis": axis}


def _find_header_and_data(rows: list[list]) -> tuple[Optional[list], Optional[list], int]:
    """Find the code row, label row, and first data row index.

    Returns (code_row, label_row, data_start_index).
    """
    # Look for the row containing "Date" in column A — that's the label row.
    for i, row in enumerate(rows[:15]):
        if row and row[0] is not None and str(row[0]).strip().lower() == "date":
            # Check if the row above has codes (UKC, A, B, etc.)
            code_row = None
            if i > 0:
                prev = rows[i - 1]
                # If the previous row has short alphanumeric codes in columns 1+
                non_empty = [c for c in prev[1:] if c is not None and str(c).strip()]
                if non_empty and all(len(str(c).strip()) <= 5 for c in non_empty):
                    code_row = prev
            return code_row, row, i + 1
    return None, None, 0


def parse_workbook(path: Path) -> list[CompensationObservation]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[CompensationObservation] = []

    for sheet_name in wb.sheetnames:
        meta = _classify_sheet(sheet_name)
        if meta is None:
            continue

        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if len(rows) < 8:
            continue

        code_row, label_row, data_start = _find_header_and_data(rows)
        if label_row is None:
            print(f"[hmrc_paye:parse] Skipping {sheet_name}: no header row found")
            continue

        # Build column metadata: index → (code, label)
        columns: dict[int, tuple[Optional[str], str]] = {}
        for col_idx in range(1, len(label_row)):
            label = label_row[col_idx]
            if label is None or str(label).strip() == "":
                continue
            code = None
            if code_row and col_idx < len(code_row):
                code = str(code_row[col_idx]).strip() if code_row[col_idx] else None
            columns[col_idx] = (code, str(label).strip())

        if not columns:
            continue

        # Parse data rows
        for row in rows[data_start:]:
            if not row or row[0] is None:
                continue
            obs_date = _parse_date_cell(row[0])
            if obs_date is None:
                continue
            # Only keep 2024+ to match our partition range
            if obs_date.year < 2024:
                continue

            for col_idx, (code, label) in columns.items():
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                if raw is None or raw == "" or raw == ".." or raw == "x":
                    continue
                try:
                    value = float(raw)
                except (TypeError, ValueError):
                    continue
                if value <= 0:
                    continue

                # Monthly pay → annualised
                normalized = value * 12

                # Determine observation type
                is_median = meta["stat"] == "median"
                obs_type = ObservationType.PERCENTILE if is_median else ObservationType.POINT
                percentile = 50 if is_median else None

                # Use NUTS code for location if available, otherwise label
                location = code if code and meta["axis"] in ("nuts1", "nuts2", "nuts3", "local_authority") else None

                ref = f"{path.stem}:{sheet_name}:{code or label}:{obs_date.isoformat()}"

                out.append(CompensationObservation(
                    source_id="hmrc_paye_rti",
                    source_reference=ref,
                    occupation_code=code if meta["axis"] == "industry" else None,
                    location_code=location,
                    company_ref=None,
                    observation_type=obs_type,
                    value_amount=value,
                    value_min=None,
                    value_max=None,
                    percentile=percentile,
                    period=Period.ANNUAL,
                    normalized_annual_amount=normalized,
                    normalization_method_version=NORMALIZATION_VERSION,
                    currency="GBP",
                    experience_band=ExperienceBand.UNKNOWN,
                    contract_type=ContractType.UNKNOWN,
                    sample_size=None,
                    total_comp_annual=None,
                    observed_at=obs_date,
                    source_payload={
                        "workbook": path.name,
                        "sheet": sheet_name,
                        "axis": meta["axis"],
                        "stat": meta["stat"],
                        "code": code,
                        "label": label,
                        "month": obs_date.isoformat(),
                        "monthly_pay_gbp": value,
                    },
                ))

    print(f"[hmrc_paye:parse] {path.name}: {len(out)} observations from {len([s for s in wb.sheetnames if _classify_sheet(s)])} sheets")
    return out


def parse_run_dir(run_dir: Path) -> list[CompensationObservation]:
    results: list[CompensationObservation] = []
    for xlsx in sorted(run_dir.glob("*.xlsx")):
        try:
            results.extend(parse_workbook(xlsx))
        except Exception as e:
            print(f"[hmrc_paye:parse] Skipping {xlsx.name}: {e}")
    return results


if __name__ == "__main__":
    import sys
    records = parse_run_dir(Path(sys.argv[1]))
    print(f"Parsed {len(records)} observations.")
